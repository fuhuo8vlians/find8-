"""
数据管理模块 — 资产 & 公司 CRUD、Excel 导入导出、DNS 解析
"""
import csv
import glob
import io
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import dns.resolver
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
ASSETS_FILE = os.path.join(DATA_DIR, "assets.json")
COMPANIES_FILE = os.path.join(DATA_DIR, "companies.json")
PROJECTS_FILE = os.path.join(DATA_DIR, "projects.json")

DNS_TIMEOUT = 10
DNS_NAMESERVERS = ["8.8.8.8", "8.8.4.4"]
MAX_WORKERS = 50

# CDN 检测数据（复用 OneForAll 的 CDN 指纹库）
_CDN_CNAME_KEYWORDS = {}
_CDN_IP_CIDR = []
_CDN_DATA_LOADED = False


def _load_cdn_data():
    global _CDN_CNAME_KEYWORDS, _CDN_IP_CIDR, _CDN_DATA_LOADED
    if _CDN_DATA_LOADED:
        return
    # OneForAll CDN 数据路径
    ofa_data = os.path.join(os.path.dirname(BASE_DIR), "子域名收集", "OneForAll-0.4.5", "data")
    try:
        cname_file = os.path.join(ofa_data, "cdn_cname_keywords.json")
        if os.path.exists(cname_file):
            with open(cname_file, "r", encoding="utf-8") as f:
                _CDN_CNAME_KEYWORDS = json.load(f)
        cidr_file = os.path.join(ofa_data, "cdn_ip_cidr.json")
        if os.path.exists(cidr_file):
            with open(cidr_file, "r", encoding="utf-8") as f:
                _CDN_IP_CIDR = json.load(f)
    except Exception:
        pass
    _CDN_DATA_LOADED = True


def _check_cdn(ip_str, cname_list):
    """检查是否为 CDN，返回 CDN 厂商名或 None"""
    _load_cdn_data()
    # 方法1: CNAME 关键词匹配
    if cname_list:
        # 额外补充 CDN 关键词（覆盖 OneForAll 未收录的）
        extra_keywords = {
            "cloudfront.net": "AWS CloudFront",
            "aliyun.com": "Aliyun CDN",
            "myqcloud.com": "Tencent CDN",
            "qiniucdn.com": "Qiniu CDN",
            "cdn.dnsv1.com": "DNSPod CDN",
            "lxdns.com": "Aliyun CDN",
        }
        for cname in cname_list:
            cname_lower = cname.lower().rstrip(".")
            for keyword, provider in extra_keywords.items():
                if keyword in cname_lower:
                    return provider
            for keyword, provider in _CDN_CNAME_KEYWORDS.items():
                if keyword in cname_lower:
                    return provider
    # 方法2: IP CIDR 匹配
    if ip_str and _CDN_IP_CIDR:
        try:
            import ipaddress
            for cidr in _CDN_IP_CIDR:
                try:
                    if ipaddress.ip_address(ip_str) in ipaddress.ip_network(cidr):
                        return "CDN"
                except Exception:
                    continue
        except Exception:
            pass
    return None


def init():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(ASSETS_FILE):
        _save_json(ASSETS_FILE, [])
    if not os.path.exists(COMPANIES_FILE):
        _save_json(COMPANIES_FILE, [{"id": 1, "name": "默认公司", "rootDomains": [], "projectId": None}])
    if not os.path.exists(PROJECTS_FILE):
        _save_json(PROJECTS_FILE, [])


def _load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ============================================================
#  项目管理
# ============================================================
def get_projects():
    projects = _load_json(PROJECTS_FILE)
    for p in projects:
        p["companyCount"] = len(get_project_companies(p["id"]))
    return projects


def get_project(project_id):
    projects = _load_json(PROJECTS_FILE)
    return next((p for p in projects if p["id"] == project_id), None)


def add_project(name, description=""):
    projects = _load_json(PROJECTS_FILE)
    new_id = max((p["id"] for p in projects), default=0) + 1
    project = {"id": new_id, "name": name, "description": description, "createdAt": datetime.now().isoformat()}
    projects.append(project)
    _save_json(PROJECTS_FILE, projects)
    return project


def delete_project(project_id):
    projects = _load_json(PROJECTS_FILE)
    projects = [p for p in projects if p["id"] != project_id]
    _save_json(PROJECTS_FILE, projects)
    # 解除公司关联
    companies = _load_json(COMPANIES_FILE)
    for c in companies:
        if c.get("projectId") == project_id:
            c["projectId"] = None
    _save_json(COMPANIES_FILE, companies)


def get_project_companies(project_id):
    companies = _load_json(COMPANIES_FILE)
    return [c for c in companies if c.get("projectId") == project_id]


def add_company_to_project(project_id, company_id):
    companies = _load_json(COMPANIES_FILE)
    company = next((c for c in companies if c["id"] == company_id), None)
    if not company:
        return None
    company["projectId"] = project_id
    _save_json(COMPANIES_FILE, companies)
    return company


def remove_company_from_project(project_id, company_id):
    companies = _load_json(COMPANIES_FILE)
    company = next((c for c in companies if c["id"] == company_id and c.get("projectId") == project_id), None)
    if company:
        company["projectId"] = None
        _save_json(COMPANIES_FILE, companies)
    return company


# ============================================================
#  公司管理
# ============================================================
def get_companies():
    return _load_json(COMPANIES_FILE)


def add_company(name, project_id=None):
    companies = _load_json(COMPANIES_FILE)
    new_id = max((c["id"] for c in companies), default=0) + 1
    company = {"id": new_id, "name": name, "rootDomains": [], "projectId": project_id}
    companies.append(company)
    _save_json(COMPANIES_FILE, companies)
    return company


def update_company_root_domains(company_id, root_domain):
    companies = _load_json(COMPANIES_FILE)
    for c in companies:
        if c["id"] == company_id:
            if root_domain not in c["rootDomains"]:
                c["rootDomains"].append(root_domain)
            break
    _save_json(COMPANIES_FILE, companies)


# ============================================================
#  资产管理
# ============================================================
def get_assets(sort_by=None, sort_order="asc", company_filter=None):
    assets = _load_json(ASSETS_FILE)
    if company_filter:
        assets = [a for a in assets if a.get("companyName", "") == company_filter]
    if sort_by:
        reverse = sort_order == "desc"
        if sort_by == "ip":
            assets.sort(key=lambda x: _ip_sort_key(x.get("ip")), reverse=reverse)
        elif sort_by == "url":
            assets.sort(key=lambda x: x.get("url", "") or "", reverse=reverse)
        elif sort_by == "companyName":
            assets.sort(key=lambda x: x.get("companyName", ""), reverse=reverse)
        elif sort_by == "rootDomain":
            assets.sort(key=lambda x: x.get("rootDomain", ""), reverse=reverse)
        elif sort_by == "subDomain":
            assets.sort(key=lambda x: x.get("subDomain", ""), reverse=reverse)
        elif sort_by == "dnsStatus":
            assets.sort(key=lambda x: x.get("dnsStatus", ""), reverse=reverse)
        elif sort_by == "timestamp":
            assets.sort(key=lambda x: x.get("timestamp", ""), reverse=reverse)
        else:
            assets.sort(key=lambda x: x.get("timestamp", ""), reverse=reverse)
    else:
        assets.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return assets


def _ip_sort_key(ip):
    if not ip:
        return (0, 0, 0, 0)
    try:
        parts = [int(p) for p in ip.split(".")]
        return tuple(parts) if len(parts) == 4 else (0, 0, 0, 0)
    except Exception:
        return (0, 0, 0, 0)


def delete_asset(asset_id):
    assets = _load_json(ASSETS_FILE)
    assets = [a for a in assets if a["id"] != asset_id]
    _save_json(ASSETS_FILE, assets)


def delete_assets_batch(asset_ids):
    assets = _load_json(ASSETS_FILE)
    ids_set = set(asset_ids)
    assets = [a for a in assets if a["id"] not in ids_set]
    _save_json(ASSETS_FILE, assets)
    return len(ids_set)


def get_company_names():
    """Return list of unique company names from assets"""
    companies = _load_json(COMPANIES_FILE)
    return sorted([c["name"] for c in companies])


def add_asset(company_name, root_domains, sub_domain="", url="", project_id=None):
    companies = _load_json(COMPANIES_FILE)
    company = next((c for c in companies if c["name"] == company_name), None)
    if not company:
        company = add_company(company_name, project_id)
        companies = _load_json(COMPANIES_FILE)
        company = next((c for c in companies if c["name"] == company_name), company)
    elif project_id is not None:
        company["projectId"] = project_id
        _save_json(COMPANIES_FILE, companies)

    for rd in root_domains:
        if rd not in company["rootDomains"]:
            company["rootDomains"].append(rd)
    _save_json(COMPANIES_FILE, companies)

    assets = _load_json(ASSETS_FILE)
    new_id = max((a["id"] for a in assets), default=0) + 1
    added = 0
    for rd in root_domains:
        normalized = normalize_root_domain(rd)
        exists = any(
            a["companyId"] == company["id"]
            and a["rootDomain"] == normalized
            and a.get("subDomain", "") == sub_domain
            for a in assets
        )
        if not exists:
            assets.append({
                "id": new_id + added,
                "companyId": company["id"],
                "companyName": company["name"],
                "rootDomain": normalized,
                "subDomain": sub_domain,
                "url": url,
                "ip": None,
                "dnsStatus": "pending",
                "statusCode": None,
                "title": None,
                "cdn": None,
                "timestamp": datetime.now().isoformat(),
            })
            added += 1

    _save_json(ASSETS_FILE, assets)
    return {"added": added}


def add_batch_assets(batch_items, on_progress=None):
    """批量添加资产 + DNS解析。on_progress(percent, message)"""
    companies = _load_json(COMPANIES_FILE)
    assets = _load_json(ASSETS_FILE)
    next_id = max((a["id"] for a in assets), default=0) + 1

    root_to_company = {}
    for c in companies:
        for rd in c.get("rootDomains", []):
            root_to_company[rd] = (c["id"], c["name"])

    to_resolve = []
    for item in batch_items:
        sub = (item.get("subDomain") or "").strip().lower()
        if not sub:
            continue
        matched_root = None
        for root in sorted(root_to_company.keys(), key=len, reverse=True):
            if _is_subdomain_of(sub, root):
                matched_root = root
                break
        if matched_root:
            cid, cname = root_to_company[matched_root]
        else:
            cid, cname = 1, "默认公司"

        exists = any(
            a["companyId"] == cid and a["rootDomain"] == matched_root and a.get("subDomain", "") == sub
            for a in assets
        )
        if not exists:
            asset = {
                "id": next_id,
                "companyId": cid,
                "companyName": cname,
                "rootDomain": matched_root or sub,
                "subDomain": sub,
                "url": item.get("url", ""),
                "ip": item.get("ip"),
                "dnsStatus": "pending",
                "statusCode": item.get("statusCode"),
                "title": item.get("title"),
                "cdn": None,
                "timestamp": datetime.now().isoformat(),
            }
            assets.append(asset)
            to_resolve.append(asset)
            next_id += 1

    _save_json(ASSETS_FILE, assets)
    total = len(to_resolve)

    if total > 0:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(_resolve_one, a): a for a in to_resolve}
            done = 0
            for future in as_completed(futures):
                asset = futures[future]
                try:
                    records = future.result()
                    if records.get("A"):
                        asset["ip"] = records["A"][0]
                    asset["dnsStatus"] = "resolved" if (records.get("A") or records.get("CNAME")) else "pending"
                    if records.get("cdn"):
                        asset["cdn"] = records["cdn"]
                except Exception:
                    asset["dnsStatus"] = "pending"
                done += 1
                if done % 50 == 0:
                    _save_json(ASSETS_FILE, assets)
                if on_progress:
                    on_progress(int(done / total * 100), f"DNS解析: {done}/{total}")

        _save_json(ASSETS_FILE, assets)

    return {"total": len(assets), "new": len(to_resolve), "resolved": total}


def update_asset_dns(asset_id):
    assets = _load_json(ASSETS_FILE)
    asset = next((a for a in assets if a["id"] == asset_id), None)
    if not asset:
        return None
    sub = asset.get("subDomain", "") or asset.get("rootDomain", "")
    try:
        records = _resolve_one({"subDomain": sub})
        if records.get("A"):
            asset["ip"] = records["A"][0]
        asset["dnsStatus"] = "resolved" if (records.get("A") or records.get("CNAME")) else "pending"
        if records.get("cdn"):
            asset["cdn"] = records["cdn"]
    except Exception:
        asset["dnsStatus"] = "pending"
    _save_json(ASSETS_FILE, assets)
    return asset


# ============================================================
#  Excel 导入
# ============================================================
def parse_excel(filepath):
    df = pd.read_excel(filepath, header=None, dtype=str)
    companies = _load_json(COMPANIES_FILE)
    assets = _load_json(ASSETS_FILE)
    next_company_id = max((c["id"] for c in companies), default=0) + 1
    next_asset_id = max((a["id"] for a in assets), default=0) + 1
    existing = {(a["companyId"], a["rootDomain"], a.get("subDomain", "")) for a in assets}

    results = []
    for _, row in df.iterrows():
        company_name = str(row.get(0, "")).strip() or "默认公司"
        root_domain = normalize_root_domain(str(row.get(1, "")).strip())
        if not root_domain:
            continue

        company = next((c for c in companies if c["name"] == company_name), None)
        if not company:
            company = {"id": next_company_id, "name": company_name, "rootDomains": [root_domain]}
            companies.append(company)
            next_company_id += 1
        elif root_domain not in company["rootDomains"]:
            company["rootDomains"].append(root_domain)

        key = (company["id"], root_domain, "")
        if key not in existing:
            results.append({
                "id": next_asset_id,
                "companyId": company["id"],
                "companyName": company_name,
                "rootDomain": root_domain,
                "subDomain": "",
                "url": "",
                "ip": None,
                "dnsStatus": "pending",
                "timestamp": datetime.now().isoformat(),
            })
            existing.add(key)
            next_asset_id += 1

    _save_json(COMPANIES_FILE, companies)

    # Save preview assets too
    assets.extend(results)
    _save_json(ASSETS_FILE, assets)

    return {"parsed": results, "total": len(results)}


# ============================================================
#  Excel 导出
# ============================================================
def export_excel_bytes(company_filter=None):
    assets = _load_json(ASSETS_FILE)
    if company_filter:
        assets = [a for a in assets if a.get("companyName", "") == company_filter]
    if not assets:
        return None

    df = pd.DataFrame(assets)
    column_mapping = {
        "id": "ID", "companyId": "公司ID", "companyName": "公司名称",
        "rootDomain": "根域名", "subDomain": "子域名", "url": "URL",
        "ip": "IP地址", "dnsStatus": "DNS状态", "timestamp": "时间戳",
    }
    df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
    if "时间戳" in df.columns:
        df["时间戳"] = pd.to_datetime(df["时间戳"]).dt.strftime("%Y-%m-%d %H:%M:%S")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="资产数据", index=False)

    output.seek(0)
    wb = load_workbook(output)
    _beautify(wb)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _beautify(wb):
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    data_font = Font(color="000000", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# ============================================================
#  DNS 解析
# ============================================================
def _create_resolver():
    resolver = dns.resolver.Resolver()
    resolver.timeout = DNS_TIMEOUT
    resolver.lifetime = DNS_TIMEOUT
    resolver.nameservers = DNS_NAMESERVERS
    return resolver


def _resolve_one(asset):
    sub = asset.get("subDomain", "") or asset.get("rootDomain", "")
    records = {"A": [], "AAAA": [], "CNAME": [], "cdn": None}
    resolver = _create_resolver()
    cnames = []
    ips = []
    for rtype in ("A", "AAAA", "CNAME"):
        try:
            answers = resolver.resolve(sub, rtype)
            records[rtype] = [str(a) for a in answers]
            if rtype == "CNAME":
                cnames = [str(a).rstrip(".") for a in answers]
            if rtype == "A":
                ips = [str(a) for a in answers]
        except Exception:
            pass
    # CDN 检测
    primary_ip = ips[0] if ips else None
    records["cdn"] = _check_cdn(primary_ip, cnames)
    return records


def resolve_stream(subdomain):
    """流式 DNS 解析生成器"""
    records = _resolve_one({"subDomain": subdomain})
    yield {"status": "started", "subdomain": subdomain}
    yield {"status": "records", "subdomain": subdomain, "records": records}
    yield {"status": "completed", "subdomain": subdomain, "ip": records["A"][0] if records["A"] else None}


def batch_resolve_stream(assets_list, on_progress=None):
    """批量流式 DNS 解析"""
    total = len(assets_list)
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_resolve_one, a): a for a in assets_list}
        done = 0
        for future in as_completed(futures):
            asset = futures[future]
            try:
                records = future.result()
                asset["ip"] = records["A"][0] if records["A"] else None
                asset["dnsStatus"] = "resolved" if (records.get("A") or records.get("CNAME")) else "pending"
            except Exception:
                asset["dnsStatus"] = "pending"
            results.append(asset)
            done += 1
            if on_progress:
                on_progress(done, total)
    _save_json(ASSETS_FILE, _load_json(ASSETS_FILE))  # persist resolved IPs
    return results


# ============================================================
#  多条件筛选
# ============================================================
def filter_assets(project_ids=None, company_ids=None, company_names=None,
                  root_domains=None, sub_domain_pattern=None, ip_pattern=None):
    """多条件组合筛选资产，条件之间 AND 关系，同类条件 OR 关系"""
    assets = _load_json(ASSETS_FILE)
    companies = _load_json(COMPANIES_FILE)

    # 按项目筛选：找出归属项目的公司 ID
    if project_ids:
        pid_set = set(project_ids)
        allowed_company_ids = set()
        for c in companies:
            if c.get("projectId") in pid_set:
                allowed_company_ids.add(c["id"])

    # 按公司ID筛选
    if company_ids:
        cid_set = set(company_ids)
        if project_ids:
            cid_set = cid_set & allowed_company_ids
        allowed_company_ids = cid_set
    elif project_ids:
        pass  # allowed_company_ids already set above
    else:
        allowed_company_ids = None

    # 按公司名筛选（OR）
    if company_names:
        name_set = set(company_names)

    result = []
    for a in assets:
        # 公司ID匹配
        if allowed_company_ids is not None and a.get("companyId") not in allowed_company_ids:
            continue
        # 公司名匹配（额外条件，AND）
        if company_names and a.get("companyName") not in name_set:
            continue
        # 根域名匹配（OR）
        if root_domains:
            matched = False
            for rd in root_domains:
                if a.get("rootDomain", "") == rd or a.get("subDomain", "").endswith("." + rd):
                    matched = True
                    break
            if not matched:
                continue
        # 子域名模糊匹配
        if sub_domain_pattern:
            pattern = sub_domain_pattern.lower()
            sub = a.get("subDomain", "").lower()
            if pattern.startswith("*") and pattern.endswith("*"):
                if pattern[1:-1] not in sub:
                    continue
            elif pattern.startswith("*"):
                if not sub.endswith(pattern[1:]):
                    continue
            elif pattern.endswith("*"):
                if not sub.startswith(pattern[:-1]):
                    continue
            else:
                if pattern not in sub:
                    continue
        # IP模糊匹配
        if ip_pattern:
            ip = a.get("ip", "") or ""
            if ip_pattern.startswith("*") and ip_pattern.endswith("*"):
                if ip_pattern[1:-1] not in ip:
                    continue
            elif ip_pattern.startswith("*"):
                if not ip.endswith(ip_pattern[1:]):
                    continue
            elif ip_pattern.endswith("*"):
                if not ip.startswith(ip_pattern[:-1]):
                    continue
            else:
                if ip_pattern not in ip:
                    continue
        result.append(a)

    result.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return result


def get_unique_values(field):
    """获取某字段的所有唯一值（用于筛选下拉框）"""
    assets = _load_json(ASSETS_FILE)
    values = set()
    for a in assets:
        v = a.get(field)
        if v and str(v).strip():
            values.add(str(v).strip())
    return sorted(values)
def normalize_root_domain(domain):
    domain = domain.lower().strip()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain


def _is_subdomain_of(subdomain, root_domain):
    s = subdomain.lower().lstrip("www.")
    r = root_domain.lower().lstrip("www.")
    return s.endswith(f".{r}") or s == r


def extract_domains(text):
    """从文本中提取有效域名"""
    domains = set()
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            if not line.startswith("http"):
                line = "http://" + line
            parsed = __import__("urllib").parse.urlparse(line)
            hostname = parsed.hostname or parsed.path.split("/")[0]
        except Exception:
            hostname = line.split("/")[0].split(":")[0]
        hostname = hostname.lower().strip()
        if hostname.startswith("www."):
            hostname = hostname[4:]
        if re.match(r"^[a-zA-Z0-9]([a-zA-Z0-9-]*[a-zA-Z0-9])?(\.[a-zA-Z0-9]([a-zA-Z0-9-]*[a-zA-Z0-9])?)+$", hostname):
            domains.add(hostname)
    return sorted(domains)
